import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename

# Função para abrir uma janela de diálogo e selecionar o arquivo
def select_file():
    Tk().withdraw()  # Evitar a criação de uma janela principal
    filename = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    return filename

# Função para abrir uma janela de diálogo e selecionar onde salvar o arquivo
def save_file():
    Tk().withdraw()
    filename = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
    return filename

# Função principal
def process_excel():
    input_file = select_file()
    if not input_file:
        print("Nenhum arquivo selecionado.")
        return

    # Carregar o arquivo Excel e mostrar as planilhas disponíveis
    workbook = load_workbook(input_file, data_only=True)
    sheets = workbook.sheetnames
    print("Planilhas disponíveis:")
    for idx, sheet in enumerate(sheets):
        print(f"{idx + 1}. {sheet}")

    # Pedir ao usuário para escolher uma planilha pelo número
    sheet_num = int(input("Digite o número da planilha que você deseja usar: "))
    if sheet_num < 1 or sheet_num > len(sheets):
        print("Número da planilha inválido.")
        return

    sheet_name = sheets[sheet_num - 1]

    # Carregar a planilha selecionada
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)

    # Identificar e guardar os valores nas células subsequentes
    operacoes_rows = df.iloc[:, 7].str.contains('operações', case=False, na=False)
    taxa_rows = df.iloc[:, 7].str.contains('Taxa', case=False, na=False)

    if operacoes_rows.any():
        operacoes_value = df.loc[operacoes_rows, df.columns[8]].values[0]
    else:
        operacoes_value = None

    if taxa_rows.any():
        taxa_value = df.loc[taxa_rows, df.columns[8]].values[0]
    else:
        taxa_value = None

    print("Valor de 'operações':", operacoes_value)
    print("Valor de 'Taxa':", taxa_value)

    # Identificar o primeiro valor na coluna A e apagar todas as linhas anteriores
    first_value_index = df[df.iloc[:, 0].notna()].index[0]

    # Usar a linha identificada como novo cabeçalho
    df.columns = df.iloc[first_value_index]
    df = df.iloc[first_value_index + 1:].reset_index(drop=True)

    #EXCLUIR COLUNAS DESNECESSÁRIAS
    df = df.drop('TipoConta', axis=1)

    #adicionar taxa e operações na ultima linha H e J Valor
    #Conta Descrição operações e taxas
    #ultima_linha = df.dropna(how='all')
    # Encontra o índice da última linha com valor
    ultima_linha = df.dropna(how='all').index[-1]
    # rever esse df.at para colocar valores em uma linha e duas colunas
    df.at[ultima_linha + 1,('Valor','ValorPag')] = operacoes_value
    df.at[ultima_linha + 2,'ValorPag'] = taxa_value
    #df.at[]

    #ordenar tabela por Conta
    df = df.sort_values("Conta")

    


    # Salvar a planilha processada em um novo arquivo
    output_file = save_file()
    if not output_file:
        print("Nenhum local de salvamento selecionado.")
        return

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("Arquivo salvo com sucesso em:", output_file)

# Executar a função principal
if __name__ == "__main__":
    process_excel()
